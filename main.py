import os
import pandas as pd
from dotenv import load_dotenv
from pathlib import Path
from categorize import categorize
from fetchers import crossref_recent_by_journal, unpaywall_lookup
from emailer import send_email

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

load_dotenv()

def extract_pub_date(item: dict) -> str:
    """Crossref JSON içinden yayın tarihini YYYY-MM-DD formatında çıkarır"""
    try:
        if "published-print" in item:
            parts = item["published-print"]["date-parts"][0]
        elif "published-online" in item:
            parts = item["published-online"]["date-parts"][0]
        elif "issued" in item:
            parts = item["issued"]["date-parts"][0]
        else:
            return ""
        return "-".join(str(x) for x in parts)
    except Exception:
        return ""

def format_excel(xlsx_path: Path):
    """Excel dosyasını düzenle: wrap text, sütun genişlikleri, URL kolonunu hyperlink yap"""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Wrap text tüm hücrelerde aktif
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Sütun genişliklerini ayarla
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)

    # URL kolonunu hyperlink yap
    url_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "url":
            url_col_idx = idx
            break

    if url_col_idx:
        for row in ws.iter_rows(min_row=2, min_col=url_col_idx, max_col=url_col_idx):
            for cell in row:
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    wb.save(xlsx_path)

def main():
    lookback = int(os.getenv("LOOKBACK_DAYS", "30"))
    max_per_journal = int(os.getenv("MAX_RESULTS_PER_JOURNAL", "0"))  # 0 = sınırsız
    unpaywall_email = os.getenv("UNPAYWALL_EMAIL", "")

    try:
        journals = pd.read_csv("journals.csv")
    except Exception as e:
        print("❌ journals.csv okunamadı:", e)
        return

    if journals.empty:
        print("❌ journals.csv boş")
        return

    print("🔹 Toplam dergi:", len(journals))

    rows = []
    for _, j in journals.iterrows():
        title = j["journal_title"]
        quartile = j["quartile"]
        journal_oa = str(j.get("open_access", "no")).lower()
        journal_oa_diamond = str(j.get("open_access_diamond", "no")).lower()

        print(f"👉 Dergi sorgulanıyor: {title}")

        try:
            items = crossref_recent_by_journal(title, days=lookback)
            print(f"   ↳ {len(items)} makale bulundu")
        except Exception as e:
            print("⚠️ Crossref hatası:", e)
            continue

        # Limit ayarı: 0 → sınırsız
        limit = None if max_per_journal == 0 else max_per_journal

        for it in items[:limit]:
            t = " ".join(it.get("title", []))
            doi = it.get("DOI", "")
            doi_url = f"https://doi.org/{doi}" if doi else ""
            pub_date = extract_pub_date(it)

            # Unpaywall OA kontrolü
            is_open, oa_status = unpaywall_lookup(doi, unpaywall_email)

            # OA sütunu: dergi OA veya diamond OA veya makale OA → YES
            if journal_oa == "yes" or journal_oa_diamond == "yes" or is_open:
                oa_value = "YES"
            else:
                oa_value = "NO"

            print(f"      • {t[:80]}... ({pub_date})")

            rows.append({
                "category": categorize(t),
                "title": t,
                "journal": title,
                "quartile": quartile,
                "pub_date": pub_date,
                "url": doi_url,
                "doi": doi,
                "OA": oa_value
            })

    if not rows:
        print("⚠️ Hiç makale bulunamadı, mail atılmayacak")
        return

    # DataFrame oluştur ve kategori + tarih sırasına göre sırala
    df = pd.DataFrame(rows).sort_values(by=["category", "pub_date"], ascending=[True, False])

    outdir = Path("output")
    outdir.mkdir(exist_ok=True)

    # Ay ismine göre dosya adı
    month_name = datetime.now().strftime("%B_%Y")  # Örn: September_2025
    xlsx_path = outdir / f"{month_name}.xlsx"

    # Excel çıktısı üret
    df.to_excel(xlsx_path, index=False, engine="openpyxl")

    # Excel dosyasını düzenle
    format_excel(xlsx_path)

    print("🔹 Excel dosyası kaydedildi:", xlsx_path)

    # Mail gönder
    send_email(
        subject=f"Endocrine Monthly Bot - {month_name} Güncel Yayınlar",
        body=f"Ekli Excel dosyasında {month_name} ayının güncel yayınları bulunmaktadır.",
        attachments=[str(xlsx_path)]
    )

if __name__ == "__main__":
    main()
